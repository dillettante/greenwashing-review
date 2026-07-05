from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .analysis import PATTERN_LABELS


NAVY, TEAL, LIGHT, BORDER = "17365D", "0F6B63", "EAF2F8", "C9D2DC"
RISK_FILL = {"매우 높음": "F4CCCC", "높음": "FCE5CD", "중간": "FFF2CC", "낮음": "D9EAD3"}

_thin = Side(style="thin", color=BORDER)
_box = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
_wrap = Alignment(wrap_text=True, vertical="top")


def _clean(value: Any) -> Any:
    if isinstance(value, str):
        return ILLEGAL_CHARACTERS_RE.sub("", value)
    return value


def _title(cell) -> None:
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.font = Font(bold=True, color="FFFFFF", size=15)
    cell.alignment = Alignment(vertical="center")


def _header_cell(cell) -> None:
    cell.fill = PatternFill("solid", fgColor=TEAL)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    cell.border = _box


def _headers(ws, headers: list[str], row: int = 1) -> None:
    for col, text in enumerate(headers, start=1):
        _header_cell(ws.cell(row=row, column=col, value=text))


def _widths(ws, widths: list[int]) -> None:
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def create_workbooks(result: dict[str, Any], authorities: list[dict[str, Any]], output_dir: Path) -> list[Path]:
    # authorities는 Node 원본과 마찬가지로 시트 내용에 직접 쓰이지 않으나 호출 호환을 위해 유지한다.
    output_dir.mkdir(parents=True, exist_ok=True)
    claims_path = output_dir / "3-claims-review.xlsx"
    evidence_path = output_dir / "3-evidence-list.xlsx"
    _build_claims(result, claims_path)
    _build_evidence(result, evidence_path)
    return [claims_path, evidence_path]


def _build_claims(a: dict[str, Any], path: Path) -> None:
    wb = Workbook()
    claim_rows = a["claims"]
    last = max(2, len(claim_rows) + 1)

    s = wb.active
    s.title = "요약"
    s.sheet_view.showGridLines = False
    s.merge_cells("A1:F1")
    s["A1"] = "그린워싱 주장 검토 요약"
    _title(s["A1"])
    info = [
        ("사건", a["matter_id"]),
        ("기업", a["context"].get("company") or "[확인 필요]"),
        ("총 주장", len(claim_rows)),
        ("매우 높음", None),
        ("높음", None),
        ("잠정평가", None),
    ]
    for offset, (key, value) in enumerate(info):
        row = 3 + offset
        label = s.cell(row=row, column=1, value=key)
        label.fill = PatternFill("solid", fgColor=LIGHT)
        label.font = Font(bold=True)
        label.border = _box
        cell = s.cell(row=row, column=2, value=_clean(value) if value is not None else None)
        cell.border = _box
    s["B6"] = f"=COUNTIF('주장별 검토'!$H$2:$H${last},\"매우 높음\")"
    s["B7"] = f"=COUNTIF('주장별 검토'!$H$2:$H${last},\"높음\")"
    s["B8"] = f"=COUNTIF('주장별 검토'!$I$2:$I${last},\"잠정평가\")"
    s.merge_cells("A10:F10")
    s["A10"] = "점수는 우선순위 도구이며 위법성의 자동 결론이 아닙니다. 최종 제출 전 변호사 검증이 필요합니다."
    s["A10"].fill = PatternFill("solid", fgColor="FFF2CC")
    s["A10"].alignment = Alignment(wrap_text=True)
    _widths(s, [18, 38, 18, 18, 18, 18])

    c = wb.create_sheet("주장별 검토")
    c.sheet_view.showGridLines = False
    headers = [
        "Claim ID", "파일", "쪽", "원문 주장", "광고성", "주장 대상", "위험점수", "위험등급",
        "평가상태", "유형", "기계 1차분류(참고)", "직접 근거 ID", "필요 증거",
        "최종 광고성", "최종 위험", "적용 조문(포섭)",
    ]
    _headers(c, headers)
    for ri, claim in enumerate(claim_rows, start=2):
        ev = claim.get("evaluation") or {}
        provisions = ", ".join(
            f"{p.get('authority_id', '')} {p.get('cite', '')}".strip() for p in ev.get("provisions", [])
        )
        values = [
            claim["claim_id"], claim["filename"], claim["page"], claim["quote"], claim["applicability"],
            claim["subject_scope"], claim["risk_score"], claim["risk_band"],
            "잠정평가" if claim["provisional"] else "자료 연결됨",
            ", ".join(PATTERN_LABELS.get(p, p) for p in claim["patterns"]),
            claim["legal_call"], ", ".join(claim["legal_basis_ids"]), "; ".join(claim["missing_evidence"]),
            ev.get("applicability_final", ""), ev.get("risk_final", ""), provisions,
        ]
        for ci, value in enumerate(values, start=1):
            cell = c.cell(row=ri, column=ci, value=_clean(value))
            cell.alignment = _wrap
        c.cell(row=ri, column=8).fill = PatternFill(
            "solid", fgColor=RISK_FILL.get(claim["risk_band"], RISK_FILL["낮음"])
        )
    c.freeze_panes = "A2"
    _widths(c, [14, 24, 7, 60, 10, 14, 10, 12, 12, 30, 42, 35, 42, 12, 12, 40])

    sc = wb.create_sheet("점수 상세")
    sc.sheet_view.showGridLines = False
    score_headers = [
        "Claim ID", "오인·허위", "실증 부족", "범위·전과정", "중요성", "절대·포괄",
        "비교 결함", "특별 가중", "확산·시정", "가중합계",
    ]
    _headers(sc, score_headers)
    score_keys = [
        "misleading_likelihood", "substantiation_gap", "scope_or_lifecycle", "consumer_materiality",
        "absolute_or_broad", "comparison_defect", "special_aggravator", "dissemination_or_remediation",
    ]
    for ri, claim in enumerate(claim_rows, start=2):
        values = [claim["claim_id"]] + [claim["component_scores"].get(k) for k in score_keys] + [claim["risk_score"]]
        for ci, value in enumerate(values, start=1):
            sc.cell(row=ri, column=ci, value=value)
    sc.freeze_panes = "A2"
    _widths(sc, [18, 14, 14, 14, 14, 14, 14, 14, 14, 14])

    wb.save(path)


def _build_evidence(a: dict[str, Any], path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "증거목록"
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:H1")
    ws["A1"] = "그린워싱 검토 증거목록"
    _title(ws["A1"])
    headers = ["증거 ID", "구분", "파일", "쪽", "SHA-256", "관련 Claim ID", "발췌·설명", "검증 상태"]
    _headers(ws, headers, row=3)

    claim_map: dict[str, list[str]] = {}
    for claim in a["claims"]:
        claim_map.setdefault(f"{claim['document_id']}:{claim['page']}", []).append(claim["claim_id"])

    seen: set[str] = set()
    ri = 4
    for page in [*a["input_documents"], *a["evidence_documents"]]:
        key = f"{page['document_id']}:{page['page']}:{page['source_type']}"
        if key in seen:
            continue
        seen.add(key)
        source_type = page["source_type"]
        evidence_id = f"{'T' if source_type == 'target' else 'E'}-{page['document_id']}-{page['page']}"
        gubun = "검토 대상" if source_type == "target" else (
            "공개 교차확인자료" if source_type == "public_evidence" else "실증·반증자료"
        )
        text = page.get("text") or ""
        excerpt = re.sub(r"\s+", " ", text)[:500] if text else "[OCR/텍스트 추출 필요]"
        if text:
            status = "공개자료·진실성 별도 검증 필요" if source_type == "public_evidence" else "해시·텍스트 추출 완료"
        else:
            status = "[확인 필요] OCR"
        values = [
            evidence_id, gubun, page["filename"], page["page"], page["sha256"],
            ", ".join(claim_map.get(f"{page['document_id']}:{page['page']}", [])), excerpt, status,
        ]
        for ci, value in enumerate(values, start=1):
            ws.cell(row=ri, column=ci, value=_clean(value)).alignment = _wrap
        ri += 1
    ws.freeze_panes = "A4"
    _widths(ws, [22, 16, 28, 7, 68, 30, 70, 24])
    wb.save(path)
